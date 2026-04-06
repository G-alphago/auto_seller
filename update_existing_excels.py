import os
import glob
from openpyxl import load_workbook

def update_existing_excels():
    # outputs 폴더 내부의 qoo10_upload_ 접두사가 붙은 모든 xlsx 파일 조회 (임시파일인 ~$ 제외)
    target_files = [f for f in glob.glob("outputs/qoo10_upload_*.xlsx") if not os.path.basename(f).startswith("~$")]
    
    if not target_files:
        print("업데이트할 엑셀 파일이 없습니다.")
        return

    print(f"총 {len(target_files)}개의 파일을 업데이트합니다...\n")

    for filepath in target_files:
        filename = os.path.basename(filepath)
        
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # 파일명 마지막 9자리 추출 (예: qoo10_upload_20260403_123629 -> 03_123629)
            base_name = os.path.splitext(filename)[0]
            file_suffix = base_name[-9:] if len(base_name) >= 9 else base_name
            
            start_row = 5
            updated_count = 0
            
            for idx, row in enumerate(range(start_row, ws.max_row + 1)):
                # 상품 번호(item_number) 등 필수값이 없는 빈 행은 데이터 끝으로 간주
                if not ws.cell(row=row, column=1).value and not ws.cell(row=row, column=3).value:
                    break
                    
                seller_code = f"{file_suffix}_{idx + 1:02d}"
                ws.cell(row=row, column=2, value=seller_code)
                updated_count += 1
                
            wb.save(filepath)
            print(f"[완료] {filename} -> {updated_count}개 상품코드 적용 (예: {file_suffix}_01)")
            
        except Exception as e:
            print(f"[오류] {filename} 처리 실패: {e}")

if __name__ == "__main__":
    update_existing_excels()
