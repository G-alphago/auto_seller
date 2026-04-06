import os
import glob
from openpyxl import load_workbook

def sync_item_codes(dry_run=True):
    # 1. outputs 폴더 내 엑셀 파일들에서 (상품명 -> 판매자상품코드) 매핑 생성
    mapping = {}
    upload_files = glob.glob("outputs/qoo10_upload_*.xlsx")
    
    for up_file in upload_files:
        if os.path.basename(up_file).startswith("~$"):
            continue
        try:
            wb = load_workbook(up_file, data_only=True)
            ws = wb.active
            
            # 보통 5번째 또는 2번째 행부터 데이터 시작
            for row in range(1, ws.max_row + 1):
                item_name = ws.cell(row=row, column=5).value # E열
                seller_code = ws.cell(row=row, column=2).value # B열
                
                # 유효한 상품명과 판매자상품코드가 있을 경우에만
                if item_name and seller_code and item_name not in ["item_name", "상품명", "필수입력"]:
                    mapping[str(item_name).strip()] = seller_code
        except Exception as e:
            print(f"[오류] 파일 로드 실패 ({up_file}): {e}")

    print(f"✅ 기존 엑셀 파일에서 총 {len(mapping)}개의 상품 코드를 찾았습니다.")
    
    # 2. 수정/Qoo10_ItemInfo_*.xlsx 파일들에 적용
    info_files = glob.glob("수정/Qoo10_ItemInfo_*.xlsx")
    
    if not info_files:
        print("❌ 수정 폴더에 Qoo10_ItemInfo_ 파일이 없습니다.")
        return

    for info_file in info_files:
        if os.path.basename(info_file).startswith("~$"):
            continue
            
        print(f"\n▶ 대조 파일: {os.path.basename(info_file)}")
        updated_count = 0
        try:
            wb = load_workbook(info_file)
            ws = wb.active
            
            for row in range(4, ws.max_row + 1): # 4행부터 데이터라 가정
                item_name = str(ws.cell(row=row, column=5).value or "").strip()
                
                if not item_name:
                    continue
                    
                if item_name in mapping:
                    target_code = mapping[item_name]
                    current_code = ws.cell(row=row, column=2).value
                    
                    if current_code != target_code:
                        if dry_run:
                            print(f"[샌드박스: 변경예정] 행 {row} - '{item_name[:15]}...': {current_code} -> {target_code}")
                            updated_count += 1
                        else:
                            ws.cell(row=row, column=2, value=target_code)
                            updated_count += 1
                
            if not dry_run and updated_count > 0:
                wb.save(info_file)
                print(f"✅ {updated_count}개의 상품 코드를 실제 파일에 성공적으로 덮어썼습니다.")
            elif dry_run:
                print(f"👀 (샌드박스 테스트) 총 {updated_count}개의 상품 코드가 업데이트 될 예정입니다.")
            else:
                print("변경 사항이 없어 아무 내용도 저장되지 않았습니다.")
                
        except Exception as e:
            print(f"❌ 작업 중 오류 발생 ({info_file}): {e}")

if __name__ == "__main__":
    import sys
    is_dry_run = "--save" not in sys.argv
    sync_item_codes(dry_run=is_dry_run)
