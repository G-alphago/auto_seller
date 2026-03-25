import openpyxl

wb = openpyxl.load_workbook('outputs/qoo10_upload_20260326_003621.xlsx')
ws = wb.active

columns_to_check = {
    'item_name': None,
    'brand_number': None,
    'category_number': None,
    'item_weight': None,
    'start_date': None,
    'end_date': None,
    'expiration_date_EXP': None
}

# 템플릿이니까 헤더가 1줄일 수도 있고 더 밑에일 수도 있음. 
# 콜럼 인덱스 찾기
header_row = 1
for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    if 'item_name' in row:
        for col_idx, val in enumerate(row):
            if val in columns_to_check:
                columns_to_check[val] = col_idx
        break
    header_row += 1

print(f"Header found at row {header_row}. Indices: {columns_to_check}")

for i, row in enumerate(ws.iter_rows(min_row=5, max_row=7, values_only=True)):
    print(f"\n--- Row {i+1} ---")
    for col_name, col_idx in columns_to_check.items():
        if col_idx is not None and col_idx < len(row):
            print(f"{col_name}: {row[col_idx]}")

