import os
import glob
import time
import openpyxl
import shutil
from datetime import datetime
from scraper import extract_product, extract_with_playwright
from converter import calculate_price_jpy
import re

def extract_numbers(price_str):
    """가격 문자열에서 숫자 부분만 추출 (예: '10,000원' -> 10000)"""
    if not price_str: return 0
    # 문자열에서 숫자만 찾아 합침
    nums = re.findall(r'\d+', str(price_str))
    if not nums: return 0
    return int(''.join(nums))

def build_baseline_data(outputs_dir):
    """모든 큐텐 업로드 파일들을 읽어 기준(old) 가격과 옵션 데이터를 메모리에 매핑"""
    baseline_data = {}
    upload_files = glob.glob(os.path.join(outputs_dir, "qoo10_upload_*.xlsx"))
    
    print(f"[로딩 시작] 총 {len(upload_files)}개의 큐텐 업로드 파일에서 기준 데이터를 추출합니다...")
    for up_file in upload_files:
        if os.path.basename(up_file).startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(up_file, data_only=True)
            ws = wb.active
            # 보통 2번째 또는 5번째 행부터 데이터가 시작됩니다.
            for row in range(1, ws.max_row + 1):
                # 업로드 양식: B열(2) 판매자상품코드, J열(10) 가격, N열(14) 옵션
                seller_code = str(ws.cell(row=row, column=2).value or "").strip()
                price_yen = ws.cell(row=row, column=10).value
                option_info = ws.cell(row=row, column=14).value
                
                if seller_code and seller_code != "seller_unique_item_id" and seller_code != "판매자상품코드":
                    baseline_data[seller_code] = {
                        "price": price_yen,
                        "options": option_info
                    }
        except Exception as e:
            print(f"  -> 파일 로드 실패 ({os.path.basename(up_file)}): {e}")

    print(f"[로딩 완료] 총 {len(baseline_data)}건의 기준 데이터 매핑 완료")
    return baseline_data


def monitor_stock():
    # 1. 환경 설정 및 기준 데이터 준비
    outputs_dir = os.path.join(os.path.dirname(__file__), "outputs")
    edit_template = "수정/Qoo10_EditItemPriceQtyList.xlsx"
    
    # 큐텐 업로드 파일들로부터 기준 데이터 생성
    baseline_data = build_baseline_data(outputs_dir)
    
    summary_files = glob.glob(os.path.join(outputs_dir, "summary_*.xlsx"))
    if not summary_files:
        print("모니터링 대상 파일(summary_*.xlsx)을 찾을 수 없습니다.")
        return

    all_changes = [] # 수정 파일에 들어갈 변경 리스트

    print(f"총 {len(summary_files)}개의 서머리 파일을 거치며 모니터링합니다.")

    for file_path in summary_files:
        if os.path.basename(file_path).startswith("~$"):
            continue
            
        print(f"\n[작업 시작] 서머리 파일: {os.path.basename(file_path)}")
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # F1에 '비고' 헤더 추가
            if ws.cell(row=1, column=6).value != "비고":
                ws.cell(row=1, column=6).value = "비고"

            # 데이터 행 (2번째 행부터)
            for row in range(2, ws.max_row + 1):
                seller_code = str(ws.cell(row=row, column=1).value or "").strip()
                item_name = str(ws.cell(row=row, column=2).value or "").strip()
                url = ws.cell(row=row, column=3).value

                if not url or "http" not in str(url):
                    continue
                    
                # [새 로직] 서머리 파일이 아닌 이전 큐텐 업로드 파일 기준 데이터를 사용합니다.
                baseline = baseline_data.get(seller_code, {})
                old_price_str = baseline.get("price", "")
                old_options_str = baseline.get("options", "")

                print(f"[{row-1}/{ws.max_row-1}] {url} 스크래핑 시도...")
                try:
                    product_data = extract_product(url)
                    
                    if not product_data.get("price") and not product_data.get("title"):
                        product_data = extract_with_playwright(url)

                    time.sleep(1) # 차단 방지

                    new_price_str = product_data.get("price", "")
                    new_options = product_data.get("options", [])
                    
                    old_price_num = extract_numbers(old_price_str)
                    new_price_krw = extract_numbers(new_price_str)
                    new_price_num = calculate_price_jpy(new_price_krw) if new_price_krw > 0 else 0

                    remark = ""
                    is_sold_out = False

                    # 1. 품절 판단
                    if not new_price_str and not product_data.get("title"):
                        remark = "품절"
                        is_sold_out = True
                    else:
                        # 2. 가격 비교
                        if old_price_num > 0 and new_price_num > 0 and old_price_num != new_price_num:
                            remark = f"가격 변동: {old_price_num}엔 -> {new_price_num}엔"
                            # 가격이 변했으므로 서머리 파일의 가격(D열=4)도 최신화
                            ws.cell(row=row, column=4).value = f"{new_price_num}엔"
                        
                        # 일부 옵션 품절 판단
                        if old_options_str and len(new_options) == 0:
                            # remark = "옵션 정보 변경 또는 품절"
                            pass
                        
                    if not remark and new_price_num == 0:
                        remark = "판매 중지 또는 가격 정보 없음(품절 의심)"
                        is_sold_out = True

                    # 변동 사항이 발생한 경우만 수집
                    if remark:
                        ws.cell(row=row, column=6).value = remark # 비고란은 6열(F열)
                        print(f"  -> [변동 감지] {remark}")
                        
                        # 변동 리스트 업
                        all_changes.append({
                            "code": seller_code,
                            "price": new_price_num,
                            "qty": 0 if is_sold_out else 50
                        })
                    else:
                        ws.cell(row=row, column=6).value = "정상"

                except Exception as e:
                    print(f"  -> 오류 발생: {e}")
                    ws.cell(row=row, column=6).value = f"에러: {e}"

            # 서머리 파일 저장 (비고 및 가격 최신화분)
            wb.save(file_path)

        except Exception as e:
            print(f"[엑셀 오류] {file_path} 처리 중 문제 발생: {e}")

    # 2. 수정용 엑셀 파일 생성 (변동 사항이 있을 때만)
    if all_changes and os.path.exists(edit_template):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_edit_file = f"수정/Qoo10_EditItemPriceQtyList_{timestamp}.xlsx"
        shutil.copy(edit_template, new_edit_file)
        
        try:
            edit_wb = openpyxl.load_workbook(new_edit_file)
            edit_ws = edit_wb.active
            
            # 사용자 요청: 5행(Row 5)부터 작성
            start_row = 5
            for i, change in enumerate(all_changes):
                current_row = start_row + i
                # 판매자상품코드는 B열(column 2)에 입력
                edit_ws.cell(row=current_row, column=2).value = change["code"]
                edit_ws.cell(row=current_row, column=3).value = "g"
                edit_ws.cell(row=current_row, column=5).value = change["price"]
                edit_ws.cell(row=current_row, column=6).value = change["qty"]
            
            edit_wb.save(new_edit_file)
            print(f"\n[수정 파일 생성 완료] {new_edit_file} (총 {len(all_changes)}건)")
        except Exception as e:
            print(f"[수정 파일 생성 실패] {e}")
    else:
        print("\n[알림] 변동 사항이 없거나 템플릿 파일이 없어 수정 파일을 생성하지 않았습니다.")

    print("\n[모든 모니터링 작업 완료]")

if __name__ == "__main__":
    monitor_stock()
