import os
import glob
import openpyxl

def fill_item_codes(dry_run=True):
    sub_dir = "수정"
    
    # 1. 아이템인포 파일 로드 및 {판매자상품코드: 상품코드} 매핑 생성
    info_files = glob.glob(os.path.join(sub_dir, "Qoo10_ItemInfo_*.xlsx"))
    info_files = [f for f in info_files if not os.path.basename(f).startswith("~$")]
    
    if not info_files:
        print("❌ 수정 폴더에 Qoo10_ItemInfo_ 파일이 없습니다.")
        return
        
    latest_info = max(info_files, key=os.path.getmtime)
    print(f"▶ 기준 아이템인포 파일: {os.path.basename(latest_info)}")
    
    code_map = {}
    try:
        wb_info = openpyxl.load_workbook(latest_info, data_only=True)
        ws_info = wb_info.active
        
        # 4행부터 데이터라 가정 (이전에 확인한 양식 기준)
        count_mapped = 0
        for row in range(4, ws_info.max_row + 1):
            item_code = str(ws_info.cell(row=row, column=1).value or "").strip() # A열: 상품코드
            seller_code = str(ws_info.cell(row=row, column=2).value or "").strip() # B열: 판매자상품코드
            
            if item_code and seller_code and seller_code not in ["판매자상품코드", "seller_unique_item_id"]:
                code_map[seller_code] = item_code
                count_mapped += 1
                
        print(f"✅ 총 {count_mapped}개의 [판매자상품코드 -> 상품코드] 매핑 구조를 만들었습니다.")
    except Exception as e:
        print(f"❌ 매핑 로드 중 오류: {e}")
        return

    # 2. 방금 만들어진 수정파일 로드 및 A열 채우기
    edit_files = glob.glob(os.path.join(sub_dir, "Qoo10_EditItemPriceQtyList_*.xlsx"))
    edit_files = [f for f in edit_files if not os.path.basename(f).startswith("~$")]
    
    if not edit_files:
         print("❌ 수정 폴더에 최근 만들어진 EditItemPriceQtyList 파일이 없습니다.")
         return
         
    # 가장 최근에 만들어진 최신 수정 파일 타겟팅
    latest_edit = max(edit_files, key=os.path.getmtime)
    print(f"\n▶ 타겟 수정 파일 (방금 생성된 파일): {os.path.basename(latest_edit)}")
    
    try:
        wb_edit = openpyxl.load_workbook(latest_edit)
        ws_edit = wb_edit.active
        
        updated_count = 0
        not_found_count = 0
        
        # 데이터는 5행부터 시작
        for row in range(5, ws_edit.max_row + 1):
            seller_code = str(ws_edit.cell(row=row, column=2).value or "").strip() # B열
            
            if not seller_code:
                continue
                
            matched_item_code = code_map.get(seller_code)
            
            if matched_item_code:
                if dry_run:
                     if updated_count < 5:
                         print(f"[샌드박스: 대조 성공] 행 {row} - 판매자코드({seller_code}) ✚ 상품코드({matched_item_code}) A열 삽입 예정")
                else:
                     ws_edit.cell(row=row, column=1, value=matched_item_code)
                updated_count += 1
            else:
                not_found_count += 1
                if dry_run and not_found_count < 3:
                     print(f"[주의] 판매자코드({seller_code})에 해당하는 상품코드를 아이템인포 파일에서 찾지 못했습니다.")
                     
        if not dry_run:
            wb_edit.save(latest_edit)
            print(f"✅ 작업 완료: {updated_count}개의 상품코드를 A열에 성공적으로 채워 넣었습니다.")
        else:
            print(f"👀 (샌드박스 결과) 총 {updated_count}개의 항목이 매칭되어 상품코드가 복사될 예정입니다. (미매칭: {not_found_count}건)")
            
    except Exception as e:
        print(f"❌ 작업 중 오류 발생 ({latest_edit}): {e}")

if __name__ == "__main__":
    import sys
    is_dry_run = "--save" not in sys.argv
    fill_item_codes(dry_run=is_dry_run)
