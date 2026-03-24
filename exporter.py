import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import re
import os
import shutil


# 0620.xlsx 컬럼 순서 (원본 유지)
COLUMNS = [
    'item_number', 'seller_unique_item_id', 'category_number', 'brand_number',
    'item_name', 'item_promotion_name', 'item_status_Y/N/D', 'start_date',
    'end_date', 'price_yen', 'retail_price_yen', 'taxrate', 'quantity',
    'option_info', 'additional_option_info', 'additional_option_text',
    'image_main_url', 'image_other_url', 'video_url', 'image_option_info',
    'image_additional_option_info', 'header_html', 'footer_html',
    'item_description', 'Shipping_number', 'option_number',
    'available_shipping_date', 'desired_shipping_date', 'search_keyword',
    'item_condition_type', 'origin_type', 'origin_region_id',
    'origin_country_id', 'origin_others', 'medication_type', 'item_weight',
    'item_material', 'model_name', 'external_product_type',
    'external_product_id', 'manufacture_date', 'expiration_date_type',
    'expiration_date_MFD', 'expiration_date_PAO', 'expiration_date_EXP',
    'under18s_display_Y/N', 'A/S_info', 'buy_limit_type',
    'buy_limit_date', 'buy_limit_qty'
]

MAX_DESC = 30000
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "0620.xlsx")


def clean_html(text: str) -> str:
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def save_to_excel(rows: list, output_dir: str = "outputs", filename: str = None) -> str:
    """큐텐 업로드용 엑셀 (여러 상품 지원)"""
    os.makedirs(output_dir, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"qoo10_upload_{timestamp}.xlsx"

    filepath = os.path.join(output_dir, filename)

    # 0620.xlsx 템플릿 복사해서 사용
    if os.path.exists(TEMPLATE_PATH):
        shutil.copy(TEMPLATE_PATH, filepath)
        wb = load_workbook(filepath)
        ws = wb.active
        # 5번째 행부터 데이터 입력
        start_row = 5
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLUMNS)
        start_row = 2

    # 데이터 입력
    for idx, row in enumerate(rows):
        current_row = start_row + idx
        
        # 상품설명 정제
        desc = clean_html(row.get("item_description", ""))
        row["item_description"] = desc[:MAX_DESC]

        # 컬럼 순서에 맞게 데이터 추출
        data_row = [row.get(col, "") for col in COLUMNS]
        
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=current_row, column=col_idx, value=value)

    wb.save(filepath)
    return filepath


def save_summary_excel(rows: list, output_dir: str = "outputs", filename: str = None) -> str:
    """상품 관리 요약 파일 생성 [상품명, URL, 가격, 옵션]"""
    os.makedirs(output_dir, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"summary_{timestamp}.xlsx"

    filepath = os.path.join(output_dir, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 헤더
    headers = ["상품명", "URL", "가격", "옵션"]
    ws.append(headers)
    
    # 데이터 (원본 데이터 기준 또는 변환 데이터 기준)
    # 여기서는 요약용이므로 가독성 좋은 데이터를 넣음
    for r in rows:
        ws.append([
            r.get("item_name", ""),
            r.get("source_url", ""), # 원본 URL을 rows에 미리 넣어둬야 함
            r.get("price_yen", ""),
            r.get("option_info", "")
        ])
        
    wb.save(filepath)
    return filepath