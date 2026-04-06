import os
import glob
from openpyxl import load_workbook

def sync_summary_codes(dry_run=True):
    # 1. 큐텐 업로드 엑셀 파일들에서 (상품명 -> 판매자상품코드) 매핑 생성
    mapping = {}
    upload_files = glob.glob("outputs/qoo10_upload_*.xlsx")
    
    for up_file in upload_files:
        if os.path.basename(up_file).startswith("~$"):
            continue
        try:
            wb_up = load_workbook(up_file, data_only=True)
            ws_up = wb_up.active
            
            for row in range(1, ws_up.max_row + 1):
                item_name = str(ws_up.cell(row=row, column=5).value or "").strip() # E열
                seller_code = str(ws_up.cell(row=row, column=2).value or "").strip() # B열
                
                if item_name and seller_code and item_name not in ["item_name", "상품명", "필수입력"]:
                    mapping[item_name] = seller_code
        except Exception as e:
            print(f"[경고] 매핑 실패 ({up_file}): {e}")

    print(f"✅ 총 {len(mapping)}개의 고유 상품 코드를 매핑했습니다.")
    
    # 2. 서머리(summary) 파일들에 적용
    summary_files = glob.glob("outputs/summary_*.xlsx")
    
    if not summary_files:
        print("❌ outputs 폴더에 summary_ 파일이 없습니다.")
        return

    for sum_file in summary_files:
        if os.path.basename(sum_file).startswith("~$"):
            continue
            
        print(f"\n▶ 처리중: {os.path.basename(sum_file)}")
        updated_count = 0
        try:
            wb = load_workbook(sum_file)
            ws = wb.active
            
            # 이미 A열이 '판매자상품코드'인지 확인하여 중복 실행 방지
            header = str(ws.cell(row=1, column=1).value or "").strip()
            
            if header == "판매자상품코드":
                print("  -> (스킵) 이미 A열에 판매자상품코드가 생성되어 있습니다.")
                continue

            # A열 앞에 새 열 삽입
            # 원래 A열(컬럼 1)이었던 데이터는 B열(컬럼 2)로 밀려남
            if not dry_run:
                ws.insert_cols(idx=1)
                ws.cell(row=1, column=1, value="판매자상품코드")
            
            # 2행부터 상품명과 대조하며 판매자상품코드 기입
            for row in range(2, ws.max_row + 1): 
                # insert_cols를 안 한 dry_run에서는 여전히 상품명이 컬럼 1이고
                # 실제 실행 시에는 컬럼 2로 밀려납니다.
                item_name_col = 1 if dry_run else 2
                code_write_col = 1 # A열
                
                item_name = str(ws.cell(row=row, column=item_name_col).value or "").strip()
                
                if not item_name:
                    continue
                    
                matched_code = mapping.get(item_name, "")
                
                if matched_code:
                    if dry_run:
                        if updated_count < 3: # 샌드박스에서는 상위 3개만 샘플 출력
                            print(f"[샌드박스: 변경예정] 행 {row} - '{item_name[:15]}...': ✚코드 {matched_code} 삽입")
                        updated_count += 1
                    else:
                        ws.cell(row=row, column=code_write_col, value=matched_code)
                        updated_count += 1
                else:
                    if not dry_run:
                        ws.cell(row=row, column=code_write_col, value="매칭실패")
                        
            if not dry_run:
                wb.save(sum_file)
                print(f"  -> ✅ 총 {updated_count}개의 상품 코드를 삽입하고 저장했습니다.")
            else:
                print(f"  👀 (샌드박스) 이 파일에서 매칭성공: {updated_count}개")
                
        except Exception as e:
            print(f"❌ 작업 중 오류 발생 ({sum_file}): {e}")

if __name__ == "__main__":
    import sys
    is_dry_run = "--save" not in sys.argv
    sync_summary_codes(dry_run=is_dry_run)
